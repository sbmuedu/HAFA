# -*- coding: utf-8 -*-
"""
تولید صفحات واژگان حافا از داده‌های «کدینگ استاندارد آموزش» (DED):
  1) فایل ایمپورت MediaWiki  (hafa_wiki_terms_import.xml) با صفحات فضای نام «واژه:»
  2) فایل نگاشت اکسل (واژگان_حافا_از_کدینگ_DED.xlsx) برای بازبینی پیش از ایمپورت

کد واژه: HAFA-EDU-TERM-### (سریال سراسری)
منبع تعریف: کدینگ استاندارد آموزش (DED) — جدول (n)
"""
import os, re, glob, html as H
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/ded_extract"
XML_OUT = "/home/user/HAFA/طراحی_ویکی_حافا/hafa_wiki_terms_import.xml"
XLSX_OUT = "/home/user/HAFA/واژگان_حافا_از_کدینگ_DED.xlsx"
NS_TERM = 3086  # فضای نام «واژه» طبق LocalSettings.php
UNIT = "واحد آمار و فناوری اطلاعات"

# ---------- ابزار استخراج (همان منطق مرحله قبل) ----------
def load_rows(fname):
    soup = BeautifulSoup(open(os.path.join(SRC, fname), encoding="utf-8", errors="replace").read(), "lxml")
    rows = []
    for t in soup.find_all("table"):
        for tr in t.find_all("tr"):
            rows.append([c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])])
    return rows


def clean(s):
    return re.sub(r"\s+", " ", s or "").strip()


def find_table_title(rows):
    for r in rows:
        if len(r) == 1 and "جدول" in r[0]:
            t = clean(r[0])
            if len(t) < 120:  # عنوان واقعی کوتاه است (متن‌های مقدمه طولانی‌اند)
                return t
    return ""


def collect_notes(rows):
    notes = []
    for r in rows:
        if len(r) == 1:
            t = clean(r[0])
            if len(t) > 40 and "جدول" not in t:
                notes.append(t)
    return notes


def parse_simple(rows, title_idx, code_idx, desc_idx=None, row_idx=1):
    recs = []
    for r in rows:
        if len(r) <= max(title_idx, code_idx):
            continue
        title = clean(r[title_idx]) if title_idx < len(r) else ""
        code = clean(r[code_idx]) if code_idx < len(r) else ""
        desc = clean(r[desc_idx]) if (desc_idx is not None and desc_idx < len(r)) else ""
        if not title and not code:
            continue
        if "کد" in code or title == "ردیف" or "ردیف" in title:
            continue
        flag = "جدید" if "جدید" in r else ""
        recs.append({"کد": code, "عنوان": title, "توضیحات": desc, "جدید": flag})
    return recs


def parse_azad(rows):
    recs = []
    for r in rows:
        if len(r) < 6:
            continue
        for ni, ci in [(1, 2), (4, 5)]:
            name = clean(r[ni]); code = clean(r[ci])
            if not name or not code or name == "نام دانشگاه":
                continue
            recs.append({"کد": code, "عنوان": name, "توضیحات": "", "جدید": ""})
    return recs


def parse_ostan(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        ko, ostan, ksh, shahr = (clean(x) for x in r[1:5])
        if not ostan and not shahr:
            continue
        if ostan == "استان":
            continue
        recs.append({"کد استان": ko, "استان": ostan, "کد شهر": ksh, "شهر": shahr})
    return recs


def parse_gerayesh(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        rid, reshte, maghta, gerayesh, code = (clean(x) for x in r[:5])
        if reshte == "عنوان رشته":
            continue
        recs.append({"رشته": reshte, "مقطع": maghta, "گرایش": gerayesh, "کد": code})
    return recs


# ---------- مشخصات جداول ----------
# (دسته, شماره جدول, فایل, نوع, برچسب مفرد, عنوان_فهرست)
CODE_TABLES = [
    ("مقاطع تحصیلی", "جدول 1", "Garde_ddf4520e04.html", "simple", "مقطع تحصیلی", 2, 3, 4),
    ("جنسیت", "جدول 13", "جنسیت_08631504e1.html", "simple", "مقدار جنسیت", 2, 3, None),
    ("وضعیت تأهل", "جدول 11", "وضعیت-تاهل_875556740a.html", "simple", "مقدار وضعیت تأهل", 2, 3, None),
    ("وضعیت دانشجو", "جدول 12", "وضعیت-دانشجو_eea442c453.html", "simple", "مقدار وضعیت دانشجو", 2, 3, None),
    ("وضعیت دوره", "جدول 14", "وضعیت-دوره_e46995bd92.html", "simple", "مقدار وضعیت دوره", 2, 3, 4),
    ("وضعیت نظام وظیفه", "جدول 4", "وضعیت-نظام-وظیفه_03a48d9e8e.html", "simple", "مقدار وضعیت نظام وظیفه", 2, 3, 4),
    ("نوع مقاطع", "جدول 15", "نوع-مقاطع_3301a51615.html", "simple", "نوع مقطع", 2, 3, None),
    ("نوع تعهد", "جدول 10", "نوع-تعهد_f954c989c5.html", "simple", "نوع تعهد", 2, 3, 4),
    ("دانشکده‌ها و پردیس‌ها", "جدول 6", "دانشکده‌-ها-و-پردیس-های-خودگردان_7ddaee6716.html", "simple", "دانشکده/پردیس", 2, 3, None),
    ("دانشگاه و موسسات مستقل و وابسته", "جدول 1-5", "دانشگاه-و-موسسات-پزشکی-مستقل-و-وابسته_98fb4ec5e3.html", "simple", "دانشگاه/دانشکده/مرکز/موسسه", 2, 3, None),
    ("دانشگاه‌های آزاد و غیرانتفاعی", "جدول 2-5", "دانشگاه‌های-آزاد-و-غیردولتی–غیرانتفاعی_d45da6185c.html", "azad", "دانشگاه آزاد/موسسه غیردولتی", None, None, None),
    ("دانشگاه‌ها و مراکز غیروابسته", "جدول 3-5", "دانشگاه‌-ها-و-مراکز-آموزشی-غیروابسته_fa1dbc727c.html", "simple", "دانشگاه/مرکز آموزشی غیروابسته", 2, 3, None),
    ("موسسات/مجتمع آموزش عالی سلامت", "جدول 7", "موسسه-و-مجتمع-آموزش-عالی-سلامت_aebccdc3de.html", "simple", "موسسه/مجتمع آموزش عالی سلامت", 2, 3, None),
    ("مجتمع آموزش عالی علوم پزشکی", "جدول 8", "مجتمع-آموزش-عالی-علوم-پزشکی_7135f96de0.html", "simple", "مجتمع آموزش عالی علوم پزشکی", 2, 3, None),
    ("موسسات طرف تعهد", "جدول 9", "موسسات-طرف-تعهد_54b201212f.html", "simple", "موسسه طرف تعهد", 2, 3, None),
    ("مراکز تحقیقاتی", "جدول 34", "مراکز-تحقیقاتی_ff2cf137f9.html", "simple", "مرکز تحقیقاتی", 2, 3, None),
    ("مراکز رشد", "جدول 35", "مراکز-رشد_189461942c.html", "simple", "مرکز رشد", 2, 3, None),
    ("پژوهشکده‌ها", "جدول 36", "پژوهشکده-ها_f3ed20cb79.html", "simple", "پژوهشکده", 2, 3, None),
    ("گرایش‌های رشته‌ها", "جدول 3", "گرایشهای-استاندارد-رشته-ها-در-مقاطع-تحصیلی_a351189a42.html", "gerayesh", "گرایش رشته", None, None, None),
    ("استان‌ها و شهرها", "جدول (استان/شهر)", "Ostan_a1d830cb84.html", "ostan", "شهر", None, None, None),
]

VALUE_LIST_CATS = {
    "مقاطع تحصیلی", "جنسیت", "وضعیت تأهل", "وضعیت دانشجو", "وضعیت دوره",
    "وضعیت نظام وظیفه", "نوع مقاطع", "نوع تعهد",
}


def norm_name(s):
    s = clean(s)
    s = re.sub(r"\s*\*+\s*$", "", s)           # حذف *** انتهایی
    s = re.sub(r"[\s♦*]+$", "", s)             # حذف ♦ و * انتهایی
    s = re.sub(r"^[\s♦*]+", "", s)
    return s


def extract(cat, tno, fname, kind, label, ti, ci, di):
    rows = load_rows(fname)
    title = find_table_title(rows)
    notes = collect_notes(rows)
    if kind == "simple":
        recs = parse_simple(rows, ti, ci, di)
    elif kind == "azad":
        recs = parse_azad(rows)
    elif kind == "ostan":
        recs = parse_ostan(rows)
    elif kind == "gerayesh":
        recs = parse_gerayesh(rows)
    else:
        recs = []
    return {"category": cat, "table_no": tno, "file": fname, "title": title,
            "notes": notes, "records": recs, "label": label, "kind": kind}


def build_terms():
    """خروجی: لیست دیکشنری واژه‌ها به ترتیب نهایی."""
    terms = []
    tables = []
    for spec in CODE_TABLES:
        cat, tno, fname, kind, label, ti, ci, di = spec
        tables.append(extract(cat, tno, fname, kind, label, ti, ci, di))

    # ۱) واژه مفهومی هر فهرست (تعریف کل فهرست)
    for d in tables:
        cat = d["category"]
        src = f"کدینگ استاندارد آموزش (DED) — {d['title'] or d['table_no']}"
        if d["notes"]:
            definition = d["notes"][0]
        else:
            definition = f"فهرست استاندارد «{cat}» در کدینگ استاندارد آموزش وزارت بهداشت، درمان و آموزش پزشکی ({d['table_no']})."
        terms.append({
            "نام واژه": norm_name(cat),
            "دسته": cat,
            "نوع": "واژه مفهومی (فهرست)",
            "کد DED": "",
            "تعریف": definition,
            "مترادف‌ها": "",
            "منبع تعریف": src,
            "وضعیت": "پیش‌نویس",
            "جدید": "",
        })

    # ۲) واژه‌های مقادیر/موجودیت‌ها
    for d in tables:
        cat = d["category"]
        label = d["label"]
        src = f"کدینگ استاندارد آموزش (DED) — {d['title'] or d['table_no']}"
        if d["kind"] == "ostan":
            for rec in d["records"]:
                city = norm_name(rec["شهر"])
                ostan = norm_name(rec["استان"])
                code = rec["کد شهر"]
                name = city
                definition = f"شهر «{city}» از استان «{ostan}» (کد استاندارد DED: {code})."
                terms.append({
                    "نام واژه": name, "دسته": cat, "نوع": "مقدار فهرست",
                    "کد DED": code, "تعریف": definition, "مترادف‌ها": "",
                    "منبع تعریف": src, "وضعیت": "پیش‌نویس", "جدید": "",
                })
        elif d["kind"] == "gerayesh":
            for rec in d["records"]:
                reshte = norm_name(rec["رشته"]); maghta = norm_name(rec["مقطع"])
                gerayesh = norm_name(rec["گرایش"]); code = rec["کد"]
                if reshte == "-" and maghta == "-" and gerayesh == "ندارد":
                    name = "ندارد"
                    definition = "عدم وجود گرایش (مقدار «ندارد») در جدول گرایش‌های استاندارد رشته‌ها (کد استاندارد DED: 0)."
                else:
                    name = f"{reshte} – {gerayesh}"
                    definition = f"گرایش «{gerayesh}» از رشته «{reshte}» در مقطع «{maghta}» (کد استاندارد DED: {code})."
                terms.append({
                    "نام واژه": name, "دسته": cat, "نوع": "گرایش",
                    "کد DED": code, "تعریف": definition, "مترادف‌ها": "",
                    "منبع تعریف": src, "وضعیت": "پیش‌نویس", "جدید": "",
                })
        else:
            for rec in d["records"]:
                name = norm_name(rec["عنوان"]); code = rec["کد"]
                desc = clean(rec["توضیحات"])
                if cat in VALUE_LIST_CATS:
                    definition = f"{label} «{name}» در کدینگ استاندارد آموزش وزارت بهداشت، درمان و آموزش پزشکی (کد استاندارد DED: {code})."
                else:
                    definition = f"{label} «{name}» با کد استاندارد {code} در کدینگ استاندارد آموزش وزارت بهداشت."
                if desc:
                    definition += " " + desc
                terms.append({
                    "نام واژه": name, "دسته": cat, "نوع": "مقدار فهرست",
                    "کد DED": code, "تعریف": definition, "مترادف‌ها": "",
                    "منبع تعریف": src, "وضعیت": "پیش‌نویس", "جدید": rec.get("جدید", ""),
                })

    # حذف تکراری‌ها (بعضی فایل‌ها دو نسخه از جدول داده دارند)
    seen = set()
    unique = []
    for t in terms:
        key = (t["دسته"], t["کد DED"], t["نام واژه"])
        if t["کد DED"] and key in seen:
            continue  # تکراری
        if t["کد DED"]:
            seen.add(key)
        unique.append(t)
    terms = unique

    # تخصیص کد سریال و تضمین یکتایی عنوان صفحه
    used_titles = {}
    for t in terms:
        base = t["نام واژه"]
        title = base
        if title in used_titles:
            used_titles[title] += 1
            title = f"{base} ({t['دسته']})"
            n = 1
            while title in used_titles:
                n += 1
                title = f"{base} ({t['دسته']} {n})"
        used_titles[title] = 1
        t["عنوان صفحه"] = title

    for i, t in enumerate(terms, start=1):
        t["کد واژه"] = f"HAFA-EDU-TERM-{i:03d}"
        t["واحد مرتبط"] = UNIT

    return terms


# ---------- ساخت wikitext ----------
def term_wikitext(t):
    return (
        "{{واژه\n"
        f"|کد واژه={t['کد واژه']}\n"
        f"|نام واژه={t['نام واژه']}\n"
        f"|تعریف={t['تعریف']}\n"
        f"|واحد مرتبط={t['واحد مرتبط']}\n"
        f"|مترادف‌ها={t['مترادف‌ها']}\n"
        f"|منبع تعریف={t['منبع تعریف']}\n"
        f"|وضعیت واژه={t['وضعیت']}\n"
        "}}\n"
    )


def esc(s):
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def build_xml(terms):
    pages = []
    for t in terms:
        text = esc(term_wikitext(t))
        page = (
            "  <page>\n"
            f"    <title>واژه:{esc(t['عنوان صفحه'])}</title>\n"
            f"    <ns>{NS_TERM}</ns>\n"
            "    <id>0</id>\n"
            "    <revision>\n"
            "      <id>0</id>\n"
            "      <timestamp>2026-08-21T00:00:00Z</timestamp>\n"
            "      <contributor>\n"
            "        <username>HafaImport</username>\n"
            "        <id>0</id>\n"
            "      </contributor>\n"
            "      <model>wikitext</model>\n"
            "      <format>text/x-wiki</format>\n"
            f"      <text xml:space=\"preserve\">{text}</text>\n"
            "    </revision>\n"
            "  </page>\n"
        )
        pages.append(page)

    header = (
        "<?xml version='1.0' encoding='utf-8'?>\n"
        "<mediawiki xmlns=\"http://www.mediawiki.org/xml/export-0.11/\" version=\"0.11\" xml:lang=\"fa\">\n"
        "  <siteinfo>\n"
        "    <sitename>حاکمیت اطلاعات و فرآیندهای آموزشی (حافا)</sitename>\n"
        "    <dbname>wiki_hafa</dbname>\n"
        "    <base>http://wiki/wiki/حافا:شروع</base>\n"
        "    <generator>MediaWiki 1.46.0</generator>\n"
        "    <case>first-letter</case>\n"
        "  </siteinfo>\n"
    )
    return header + "".join(pages) + "</mediawiki>\n"


# ---------- ساخت اکسل نگاشت ----------
FONT_TITLE = Font(name="B Nazanin", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="B Nazanin", size=11, color="1F1F1F")
FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_ALT = PatternFill("solid", fgColor="DEEBF7")
THIN = Side(style="thin", color="9BC2E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def build_xlsx(terms):
    wb = Workbook()
    ws = wb.active
    ws.title = "واژگان حافا"
    ws.sheet_view.rightToLeft = True
    widths = [18, 30, 30, 16, 16, 70, 30, 16, 12, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["کد واژه", "نام واژه", "دسته", "نوع", "کد استاندارد DED",
               "تعریف", "منبع تعریف", "وضعیت واژه", "جدید (DED)", "عنوان صفحه ویکی"]
    # تیتر
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value="واژگان حافا — استخراج‌شده از کدینگ استاندارد آموزش (DED)")
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = CENTER
    ws.row_dimensions[1].height = 26

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22

    r = 3
    for idx, t in enumerate(terms):
        vals = [t["کد واژه"], t["نام واژه"], t["دسته"], t["نوع"], t["کد DED"],
                t["تعریف"], t["منبع تعریف"], t["وضعیت"], t["جدید"], t["عنوان صفحه"]]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = FONT_BODY; c.alignment = RIGHT; c.border = BORDER
            if idx % 2 == 1:
                c.fill = FILL_ALT
        r += 1

    ws.freeze_panes = "A3"
    wb.save(XLSX_OUT)
    return len(terms)


if __name__ == "__main__":
    terms = build_terms()
    xml = build_xml(terms)
    with open(XML_OUT, "w", encoding="utf-8") as f:
        f.write(xml)
    n = build_xlsx(terms)
    print("واژه‌های تولیدشده:", n)
    print("XML:", XML_OUT)
    print("XLSX:", XLSX_OUT)
