# -*- coding: utf-8 -*-
"""
تولید بسته ایمپورت حافا از داده‌های «کدینگ استاندارد آموزش» (DED) با مدل داده تفکیک‌شده:

  واژه‌ها (فضای‌نام واژه:)      → مفاهیم + مقادیر فهرست‌های اصطلاحی (مقاطع، جنسیت، گرایش و…)
  واحدهای سازمانی (فضای‌نام اصلی) → ساختار داخلی مبنای شروع + مراجع بیرونی (دانشگاه‌ها، دانشکده‌ها، مراکز و…)
  عناصر داده (فضای‌نام داده:)    → «کد استان» و «کد شهرستان» با دامنه مقادیر DED
  صفحه مرجع (فضای‌نام حافا:)     → فهرست کامل استان‌ها و شهرستان‌ها

خروجی‌ها:
  طراحی_ویکی_حافا/hafa_wiki_terms_import.xml     (بازتولید؛ فقط واژه‌ها)
  طراحی_ویکی_حافا/hafa_wiki_orgunits_import.xml  (واحدها + سامانه DED + عناصر داده + صفحه مرجع)
  طراحی_ویکی_حافا/patch_ded_ref_model.xml        (الگو/فرم به‌روزشده + ویژگی‌ها + رده‌های جدید)
  واژگان_حافا_از_کدینگ_DED.xlsx (و HAFA_terms_from_DED.xlsx)
  واحدهای_سازمانی_حافا_از_کدینگ_DED.xlsx (و HAFA_orgunits_from_DED.xlsx)

کدگذاری: HAFA-EDU-TERM-### برای واژه‌ها، HAFA-EDU-ORG-### برای واحدهای داخلی،
بلاک ۹xx (DATA-901/902، SYS-901) برای مراجع بیرونی رزرو شده است.
"""
import os, re
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/ded_extract"
BASE = "/home/user/HAFA"
XML_TERMS = f"{BASE}/طراحی_ویکی_حافا/hafa_wiki_terms_import.xml"
XML_ORGS = f"{BASE}/طراحی_ویکی_حافا/hafa_wiki_orgunits_import.xml"
XML_PATCH = f"{BASE}/طراحی_ویکی_حافا/patch_ded_ref_model.xml"
XLSX_TERMS_FA = f"{BASE}/واژگان_حافا_از_کدینگ_DED.xlsx"
XLSX_TERMS_EN = f"{BASE}/HAFA_terms_from_DED.xlsx"
XLSX_ORGS_FA = f"{BASE}/واحدهای_سازمانی_حافا_از_کدینگ_DED.xlsx"
XLSX_ORGS_EN = f"{BASE}/HAFA_orgunits_from_DED.xlsx"
NS_TERM, NS_DATA, NS_HAFA, NS_SYS = 3086, 3088, 3092, 3084
UNIT = "واحد آمار و فناوری اطلاعات"
TS = "2026-08-21T00:00:00Z"

# ---------------------------------------------------------------- پارس DED
def load_rows(fname):
    soup = BeautifulSoup(open(os.path.join(SRC, fname), encoding="utf-8", errors="replace").read(), "lxml")
    rows = []
    for t in soup.find_all("table"):
        for tr in t.find_all("tr"):
            rows.append([c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])])
    return rows


def clean(s):
    return re.sub(r"\s+", " ", s or "").strip()


def norm_name(s):
    s = clean(s)
    s = re.sub(r"\s*\*+\s*$", "", s)
    s = re.sub(r"[\s♦*]+$", "", s)
    s = re.sub(r"^[\s♦*]+", "", s)
    return s


def find_table_title(rows):
    for r in rows:
        if len(r) == 1 and "جدول" in r[0]:
            t = clean(r[0])
            if len(t) < 120:
                return t
    return ""


def collect_notes(rows):
    return [clean(r[0]) for r in rows if len(r) == 1 and len(clean(r[0])) > 40 and "جدول" not in clean(r[0])]


def parse_simple(rows, title_idx, code_idx, desc_idx=None):
    recs = []
    for r in rows:
        if len(r) <= max(title_idx, code_idx):
            continue
        title = clean(r[title_idx]); code = clean(r[code_idx])
        desc = clean(r[desc_idx]) if (desc_idx is not None and desc_idx < len(r)) else ""
        if not title and not code:
            continue
        if "کد" in code or title == "ردیف" or "ردیف" in title:
            continue
        recs.append({"کد": code, "عنوان": title, "توضیحات": desc,
                     "جدید": "جدید" if "جدید" in r else ""})
    return recs


def parse_azad(rows):
    recs = []
    for r in rows:
        if len(r) < 6:
            continue
        for ni, ci in [(1, 2), (4, 5)]:
            name = clean(r[ni]); code = clean(r[ci])
            if not name or not code or name == "نام دانشگاه":
                continue
            recs.append({"کد": code, "عنوان": name, "توضیحات": "", "جدید": ""})
    return recs


def parse_ostan(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        ko, ostan, ksh, shahr = (clean(x) for x in r[1:5])
        if (not ostan and not shahr) or ostan == "استان":
            continue
        recs.append({"کد استان": ko, "استان": ostan, "کد شهر": ksh, "شهر": shahr})
    return recs


def parse_gerayesh(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        rid, reshte, maghta, gerayesh, code = (clean(x) for x in r[:5])
        if reshte == "عنوان رشته":
            continue
        recs.append({"رشته": reshte, "مقطع": maghta, "گرایش": gerayesh, "کد": code})
    return recs


# (دسته, شماره جدول, فایل, نوع پارس, برچسب, ایندکس عنوان, ایندکس کد, ایندکس توضیح)
CODE_TABLES = [
    ("مقاطع تحصیلی", "جدول 1", "Garde_ddf4520e04.html", "simple", "مقطع تحصیلی", 2, 3, 4),
    ("جنسیت", "جدول 13", "جنسیت_08631504e1.html", "simple", "مقدار جنسیت", 2, 3, None),
    ("وضعیت تأهل", "جدول 11", "وضعیت-تاهل_875556740a.html", "simple", "مقدار وضعیت تأهل", 2, 3, None),
    ("وضعیت دانشجو", "جدول 12", "وضعیت-دانشجو_eea442c453.html", "simple", "مقدار وضعیت دانشجو", 2, 3, None),
    ("وضعیت دوره", "جدول 14", "وضعیت-دوره_e46995bd92.html", "simple", "مقدار وضعیت دوره", 2, 3, 4),
    ("وضعیت نظام وظیفه", "جدول 4", "وضعیت-نظام-وظیفه_03a48d9e8e.html", "simple", "مقدار وضعیت نظام وظیفه", 2, 3, 4),
    ("نوع مقاطع", "جدول 15", "نوع-مقاطع_3301a51615.html", "simple", "نوع مقطع", 2, 3, None),
    ("نوع تعهد", "جدول 10", "نوع-تعهد_f954c989c5.html", "simple", "نوع تعهد", 2, 3, 4),
    ("دانشکده‌ها و پردیس‌ها", "جدول 6", "دانشکده‌-ها-و-پردیس-های-خودگردان_7ddaee6716.html", "simple", "دانشکده/پردیس", 2, 3, None),
    ("دانشگاه و موسسات مستقل و وابسته", "جدول 1-5", "دانشگاه-و-موسسات-پزشکی-مستقل-و-وابسته_98fb4ec5e3.html", "simple", "دانشگاه/موسسه", 2, 3, None),
    ("دانشگاه‌های آزاد و غیرانتفاعی", "جدول 2-5", "دانشگاه‌های-آزاد-و-غیردولتی–غیرانتفاعی_d45da6185c.html", "azad", "دانشگاه آزاد/موسسه غیردولتی", None, None, None),
    ("دانشگاه‌ها و مراکز غیروابسته", "جدول 3-5", "دانشگاه‌-ها-و-مراکز-آموزشی-غیروابسته_fa1dbc727c.html", "simple", "دانشگاه/مرکز آموزشی غیروابسته", 2, 3, None),
    ("موسسات/مجتمع آموزش عالی سلامت", "جدول 7", "موسسه-و-مجتمع-آموزش-عالی-سلامت_aebccdc3de.html", "simple", "موسسه/مجتمع آموزش عالی سلامت", 2, 3, None),
    ("مجتمع آموزش عالی علوم پزشکی", "جدول 8", "مجتمع-آموزش-عالی-علوم-پزشکی_7135f96de0.html", "simple", "مجتمع آموزش عالی علوم پزشکی", 2, 3, None),
    ("موسسات طرف تعهد", "جدول 9", "موسسات-طرف-تعهد_54b201212f.html", "simple", "موسسه طرف تعهد", 2, 3, None),
    ("مراکز تحقیقاتی", "جدول 34", "مراکز-تحقیقاتی_ff2cf137f9.html", "simple", "مرکز تحقیقاتی", 2, 3, None),
    ("مراکز رشد", "جدول 35", "مراکز-رشد_189461942c.html", "simple", "مرکز رشد", 2, 3, None),
    ("پژوهشکده‌ها", "جدول 36", "پژوهشکده-ها_f3ed20cb79.html", "simple", "پژوهشکده", 2, 3, None),
    ("گرایش‌های رشته‌ها", "جدول 3", "گرایشهای-استاندارد-رشته-ها-در-مقاطع-تحصیلی_a351189a42.html", "gerayesh", "گرایش رشته", None, None, None),
    ("استان‌ها و شهرها", "جدول (استان/شهر)", "Ostan_a1d830cb84.html", "ostan", "شهر", None, None, None),
]

# مقصد مدل برای هر دسته
VALUE_LIST_CATS = {"مقاطع تحصیلی", "جنسیت", "وضعیت تأهل", "وضعیت دانشجو", "وضعیت دوره",
                   "وضعیت نظام وظیفه", "نوع مقاطع", "نوع تعهد", "گرایش‌های رشته‌ها"}
ORG_CATS = {"دانشکده‌ها و پردیس‌ها", "دانشگاه و موسسات مستقل و وابسته", "دانشگاه‌های آزاد و غیرانتفاعی",
            "دانشگاه‌ها و مراکز غیروابسته", "موسسات/مجتمع آموزش عالی سلامت", "مجتمع آموزش عالی علوم پزشکی",
            "موسسات طرف تعهد", "مراکز تحقیقاتی", "مراکز رشد", "پژوهشکده‌ها"}
OSTAN_CAT = "استان‌ها و شهرها"


def org_type(cat, name):
    if cat == "دانشکده‌ها و پردیس‌ها":
        return "پردیس" if "پردیس" in name else "دانشکده"
    if cat in ("دانشگاه و موسسات مستقل و وابسته", "دانشگاه‌های آزاد و غیرانتفاعی", "دانشگاه‌ها و مراکز غیروابسته"):
        return "دانشگاه" if "دانشگاه" in name else "مؤسسه آموزشی"
    if cat == "موسسات/مجتمع آموزش عالی سلامت":
        return "مجتمع آموزش عالی" if "مجتمع" in name else "مؤسسه آموزشی"
    if cat == "مجتمع آموزش عالی علوم پزشکی":
        return "مجتمع آموزش عالی"
    if cat == "موسسات طرف تعهد":
        return "مؤسسه طرف تعهد"
    if cat == "مراکز تحقیقاتی":
        return "مرکز تحقیقاتی"
    if cat == "مراکز رشد":
        return "مرکز رشد"
    if cat == "پژوهشکده‌ها":
        return "پژوهشکده"
    return "واحد مرجع"


TYPE_CATEGORY = {  # نوع واحد → رده ویکی
    "دانشگاه": "رده:دانشگاه‌ها",
    "دانشکده": "رده:دانشکده‌ها و پردیس‌ها",
    "پردیس": "رده:دانشکده‌ها و پردیس‌ها",
    "مؤسسه آموزشی": "رده:مؤسسات آموزش عالی",
    "مجتمع آموزش عالی": "رده:مجتمع‌های آموزش عالی",
    "مؤسسه طرف تعهد": "رده:مؤسسات طرف تعهد",
    "مرکز تحقیقاتی": "رده:مراکز تحقیقاتی",
    "مرکز رشد": "رده:مراکز رشد",
    "پژوهشکده": "رده:پژوهشکده‌ها",
}


def extract_tables():
    tables = []
    for cat, tno, fname, kind, label, ti, ci, di in CODE_TABLES:
        rows = load_rows(fname)
        if kind == "simple":
            recs = parse_simple(rows, ti, ci, di)
        elif kind == "azad":
            recs = parse_azad(rows)
        elif kind == "ostan":
            recs = parse_ostan(rows)
        elif kind == "gerayesh":
            recs = parse_gerayesh(rows)
        else:
            recs = []
        tables.append({"category": cat, "table_no": tno, "file": fname, "kind": kind,
                       "label": label, "title": find_table_title(rows),
                       "notes": collect_notes(rows), "records": recs})
    return tables


def dedup_records(records):
    seen, out = set(), []
    for r in records:
        key = (r.get("کد", ""), r.get("عنوان", "") or r.get("شهر", ""))
        if r.get("کد") and key in seen:
            continue
        if r.get("کد"):
            seen.add(key)
        out.append(r)
    return out


# ---------------------------------------------------------------- ساخت رکوردها
def build_all(tables):
    terms, orgs, ostan_rows = [], [], []
    stats = {}

    # ۱) واژه مفهومی هر دسته (تعریف فهرست + اشاره به مقصد مدل)
    for d in tables:
        cat = d["category"]
        src = f"کدینگ استاندارد آموزش (DED) — {d['title'] or d['table_no']}"
        definition = d["notes"][0] if d["notes"] else \
            f"فهرست استاندارد «{cat}» در کدینگ استاندارد آموزش وزارت بهداشت، درمان و آموزش پزشکی ({d['table_no']})."
        if cat in ORG_CATS:
            definition += " مقادیر این فهرست نه واژه، بلکه موجودیت‌اند و در قالب «واحد سازمانی» (حوزه: مرجع بیرونی) ثبت می‌شوند."
        elif cat == OSTAN_CAT:
            definition += " مقادیر این فهرست، دامنه مقادیر مجاز عناصر داده «کد استان» و «کد شهرستان» هستند و در قالب عنصر داده ثبت می‌شوند."
        else:
            definition += " مقادیر این فهرست به‌صورت واژه در همین فضای‌نام ثبت شده‌اند."
        terms.append({"نام واژه": norm_name(cat), "دسته": cat, "نوع": "واژه مفهومی (فهرست)",
                      "کد DED": "", "تعریف": definition, "مترادف‌ها": "",
                      "منبع تعریف": src, "وضعیت": "پیش‌نویس", "جدید": "", "قالب": "واژه"})

    # ۲) رکوردهای هر دسته
    for d in tables:
        cat = d["category"]
        label, src = d["label"], f"کدینگ استاندارد آموزش (DED) — {d['title'] or d['table_no']}"

        if cat == OSTAN_CAT:
            for rec in d["records"]:
                ostan_rows.append({"کد استان": rec["کد استان"], "استان": norm_name(rec["استان"]),
                                   "کد شهر": rec["کد شهر"], "شهر": norm_name(rec["شهر"])})
            stats[cat] = len(d["records"])
            continue

        if cat in ORG_CATS:
            recs = dedup_records(d["records"])
            for rec in recs:
                name = norm_name(rec["عنوان"]); code = rec["کد"]
                orgs.append({"حوزه": "مرجع بیرونی", "کد واحد": "", "نام واحد": name,
                             "نوع واحد": org_type(cat, name), "کد DED": code,
                             "دسته": cat, "جدول": d["table_no"], "جدید": rec.get("جدید", ""),
                             "شرح": f"واحد مرجع بیرونی ثبت‌شده در کدینگ استاندارد آموزش وزارت بهداشت ({d['table_no']}، کد {code})."
                                    + (f" {clean(rec['توضیحات'])}" if clean(rec.get("توضیحات", "")) else "")})
            stats[cat] = len(recs)
            continue

        # واژه‌ها
        if d["kind"] == "gerayesh":
            recs = d["records"]
            for rec in recs:
                reshte, maghta = norm_name(rec["رشته"]), norm_name(rec["مقطع"])
                gerayesh, code = norm_name(rec["گرایش"]), rec["کد"]
                if reshte == "-" and maghta == "-" and gerayesh == "ندارد":
                    name = "ندارد"
                    definition = "عدم وجود گرایش (مقدار «ندارد») در جدول گرایش‌های استاندارد رشته‌ها (کد استاندارد DED: 0)."
                else:
                    name = f"{reshte} – {gerayesh}"
                    definition = f"گرایش «{gerayesh}» از رشته «{reshte}» در مقطع «{maghta}» (کد استاندارد DED: {code})."
                terms.append({"نام واژه": name, "دسته": cat, "نوع": "گرایش", "کد DED": code,
                              "تعریف": definition, "مترادف‌ها": "", "منبع تعریف": src,
                              "وضعیت": "پیش‌نویس", "جدید": "", "قالب": "واژه"})
        else:
            recs = dedup_records(d["records"])
            for rec in recs:
                name, code = norm_name(rec["عنوان"]), rec["کد"]
                desc = clean(rec.get("توضیحات", ""))
                definition = f"{label} «{name}» در کدینگ استاندارد آموزش وزارت بهداشت، درمان و آموزش پزشکی (کد استاندارد DED: {code})."
                if desc:
                    definition += " " + desc
                terms.append({"نام واژه": name, "دسته": cat, "نوع": "مقدار فهرست", "کد DED": code,
                              "تعریف": definition, "مترادف‌ها": "", "منبع تعریف": src,
                              "وضعیت": "پیش‌نویس", "جدید": rec.get("جدید", ""), "قالب": "واژه"})
        stats[cat] = len(recs)

    return terms, orgs, ostan_rows, stats


# ---------------------------------------------------------------- واحدهای داخلی مبنای شروع
SBMU_ROOT = "دانشگاه علوم پزشکی شهید بهشتی"
INTERNAL_UNITS = [
    ("معاونت آموزشی", "معاونت", SBMU_ROOT, "معاونت دانشگاه در حوزه آموزش؛ متولی پروژه حافا و مالک فرآیندهای آموزشی سطح دانشگاه. (پیش‌نویس مبنای شروع — در جلسه آغاز تکمیل شود.)"),
    ("مدیریت امور آموزشی", "مدیریت", "معاونت آموزشی", "مدیریت اجرای امور آموزشی مقاطع تحصیلی. (پیش‌نویس مبنای شروع.)"),
    ("مدیریت تحصیلات تکمیلی", "مدیریت", "معاونت آموزشی", "مدیریت امور تحصیلات تکمیلی (ارشد و دکتری). (پیش‌نویس مبنای شروع.)"),
    ("مرکز آزمون", "مرکز", "معاونت آموزشی", "برگزاری آزمون‌های آموزشی. (پیش‌نویس مبنای شروع.)"),
    ("مرکز مطالعات و توسعه آموزش پزشکی", "مرکز", "معاونت آموزشی", "توسعه آموزش و تحلیل‌های آموزشی. (پیش‌نویس مبنای شروع.)"),
    ("دفتر آموزش مداوم", "دفتر", "معاونت آموزشی", "آموزش مداوم و بازتوانی. (پیش‌نویس مبنای شروع.)"),
]


def build_internal(orgs):
    """واحدهای داخلی با کد HAFA-EDU-ORG؛ اگر دانشگاه در فهرست DED باشد، کد آن به ریشه منتقل و رکورد بیرونی حذف می‌شود."""
    ded_code = ""
    for o in orgs:
        if o["نام واحد"] == SBMU_ROOT and o["حوزه"] == "مرجع بیرونی":
            ded_code = o["کد DED"]
            break
    orgs = [o for o in orgs if not (o["نام واحد"] == SBMU_ROOT and o["حوزه"] == "مرجع بیرونی")]

    internal = [{"حوزه": "داخلی", "کد واحد": "HAFA-EDU-ORG-001", "نام واحد": SBMU_ROOT,
                 "نوع واحد": "دانشگاه", "کد DED": ded_code, "دسته": "ساختار داخلی (مبنای شروع)",
                 "جدول": "—", "جدید": "",
                 "شرح": "نهاد ریشه ساختار دانشگاه در مدل حافا. (پیش‌نویس مبنای شروع.)"}]
    for i, (name, typ, parent, mission) in enumerate(INTERNAL_UNITS, start=2):
        internal.append({"حوزه": "داخلی", "کد واحد": f"HAFA-EDU-ORG-{i:03d}", "نام واحد": name,
                         "نوع واحد": typ, "کد DED": "", "دسته": "ساختار داخلی (مبنای شروع)",
                         "جدول": "—", "جدید": "", "شرح": mission, "والد": parent})
    return internal


# ---------------------------------------------------------------- wikitext
def esc(s):
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def term_wikitext(t):
    return ("{{واژه\n"
            f"|کد واژه={t['کد واژه']}\n"
            f"|نام واژه={t['نام واژه']}\n"
            f"|تعریف={t['تعریف']}\n"
            f"|واحد مرتبط={UNIT}\n"
            f"|مترادف‌ها={t['مترادف‌ها']}\n"
            f"|منبع تعریف={t['منبع تعریف']}\n"
            f"|وضعیت واژه={t['وضعیت']}\n"
            "}}\n")


def org_wikitext(o):
    lines = ["{{واحد سازمانی",
             f"|کد واحد={o['کد واحد']}",
             f"|نام واحد={o['نام واحد']}",
             f"|نوع واحد={o['نوع واحد']}",
             f"|واحد بالادست={o.get('والد', '')}",
             "|مدیر واحد=",
             "|وضعیت واحد=فعال",
             f"|شرح مأموریت={o['شرح']}",
             f"|کد استاندارد DED={o['کد DED']}",
             f"|حوزه واحد={o['حوزه']}",
             "}}\n"]
    out = "\n".join(lines)
    if o["حوزه"] == "مرجع بیرونی":
        out += f"[[{TYPE_CATEGORY.get(o['نوع واحد'], 'رده:واحدهای سازمانی')}]]\n"
    return out


def sys_wikitext():
    return ("{{سامانه\n"
            "|کد سامانه=HAFA-EDU-SYS-901\n"
            "|نام سامانه=کدینگ استاندارد آموزش (DED)\n"
            "|مالک سامانه=معاونت آموزشی وزارت بهداشت، درمان و آموزش پزشکی\n"
            "|وضعیت سامانه=فعال\n"
            "|آدرس وب=http://ded.behdasht.gov.ir\n"
            "|توضیحات=مرجع رسمی کدینگ استاندارد آموزش (دانشگاه‌ها، مقاطع، رشته‌ها، مراکز و…). "
            "واژه‌ها، واحدهای سازمانی مرجع و دامنه‌های استان/شهر حافا از این منبع استخراج شده‌اند "
            "(بازدید: مرداد ۱۴۰۵). بلاک کد ۹xx برای سامانه‌های مرجع بیرونی رزرو شده است.\n"
            "}}\n")


def data_element_wikitext(code, name, definition, dtype, fmt, domain, extra_note=""):
    return ("{{عنصر داده\n"
            f"|کد عنصر داده={code}\n"
            f"|نام عنصر داده={name}\n"
            f"|تعریف عنصر داده={definition}\n"
            f"|نوع داده={dtype}\n"
            f"|قالب داده={fmt}\n"
            f"|دامنه مقادیر={domain}\n"
            f"|مالک داده=معاونت آموزشی وزارت بهداشت (کدینگ استاندارد آموزش)\n"
            "|سامانه مرجع=سامانه:کدینگ استاندارد آموزش (DED)\n"
            f"|واحد مسئول داده={UNIT}\n"
            "|سطح حساسیت=داخلی\n"
            "}}\n" + extra_note)


def ostan_ref_wikitext(rows):
    lines = ["# فهرست مرجع استان‌ها و شهرستان‌ها مطابق «کدینگ استاندارد آموزش» (DED) — منبع تغذیه عناصر داده «کد استان» و «کد شهرستان».",
             "",
             "'''منبع:''' کدینگ استاندارد آموزش معاونت آموزشی وزارت بهداشت (ded.behdasht.gov.ir) — بازدید: مرداد ۱۴۰۵. "
             "این صفحه فقط مرجع مطالعه است؛ تغییر مقادیر تنها پس از به‌روزرسانی منبع DED و تأیید کارگروه واژگان و حاکمیت داده انجام می‌شود.",
             "",
             "تعداد کل ردیف‌ها: " + str(len(rows)) + ".",
             "",
             '{| class="wikitable mw-collapsible" style="width:60%; direction:rtl;"',
             "! کد استان !! استان !! کد شهر !! شهر"]
    for r in rows:
        lines.append(f"|-\n| {r['کد استان']} || {r['استان']} || {r['کد شهر']} || {r['شهر']}")
    lines.append("|}")
    return "\n".join(lines) + "\n"


def xml_page(title, ns, text):
    return ("  <page>\n"
            f"    <title>{esc(title)}</title>\n"
            f"    <ns>{ns}</ns>\n"
            "    <id>0</id>\n"
            "    <revision>\n      <id>0</id>\n"
            f"      <timestamp>{TS}</timestamp>\n"
            "      <contributor>\n        <username>HafaImport</username>\n        <id>0</id>\n      </contributor>\n"
            "      <model>wikitext</model>\n      <format>text/x-wiki</format>\n"
            f"      <text xml:space=\"preserve\">{esc(text)}</text>\n"
            "    </revision>\n  </page>\n")


def xml_doc(pages):
    return ("<?xml version='1.0' encoding='utf-8'?>\n"
            "<mediawiki xmlns=\"http://www.mediawiki.org/xml/export-0.11/\" version=\"0.11\" xml:lang=\"fa\">\n"
            "  <siteinfo>\n"
            "    <sitename>حاکمیت اطلاعات و فرآیندهای آموزشی (حافا)</sitename>\n"
            "    <dbname>wiki_hafa</dbname>\n"
            "    <base>http://wiki/wiki/حافا:شروع</base>\n"
            "    <generator>MediaWiki 1.46.0</generator>\n"
            "    <case>first-letter</case>\n"
            "  </siteinfo>\n" + "".join(pages) + "</mediawiki>\n")


# ---------------------------------------------------------------- الگو/فرم به‌روزشده (patch)
TPL_ORG = """<noinclude>
الگوی واحد سازمانی حافا؛ برای ثبت معاونت، مدیریت، مرکز، دفتر، اداره و نیز واحدهای مرجع بیرونی (دانشگاه‌ها، دانشکده‌ها، مراکز و مؤسساتِ فهرست‌های کدینگ استاندارد آموزش) استفاده می‌شود.
</noinclude><includeonly>
<div style="direction:rtl; text-align:right;">
{| class="wikitable" style="width:100%;"
|+ '''واحد سازمانی: {{{نام واحد|{{PAGENAME}}}}}'''
|-
! کد واحد || {{{کد واحد|}}}
! نوع واحد || {{{نوع واحد|}}}
! واحد بالادست || [[{{{واحد بالادست|}}}]]
|-
! کد استاندارد DED || {{{کد استاندارد DED|—}}}
! حوزه واحد || {{{حوزه واحد|داخلی}}}
! وضعیت || {{{وضعیت واحد|فعال}}}
|-
! مدیر واحد || [[{{{مدیر واحد|}}}]]
! شرح مأموریت || colspan="5" | {{{شرح مأموریت|}}}
|}

== فرآیندهای تحت مسئولیت ==
{{#ask:[[رده:فرآیندها]][[واحد مسئول::{{FULLPAGENAME}}]]
|?کد فرآیند
|?سطح فرآیند
|?مالک فرآیند
|?وضعیت حاکمیتی
|format=table
|mainlabel=فرآیند
|sort=کد فرآیند
|default=هنوز فرآیندی برای این واحد ثبت نشده است.
}}

== فرآیندهای همکار ==
{{#ask:[[رده:فرآیندها]][[واحدهای همکار::{{FULLPAGENAME}}]]
|?کد فرآیند
|?واحد مسئول
|?وضعیت حاکمیتی
|format=table
|mainlabel=فرآیند
|sort=کد فرآیند
|default=این واحد در فرآیند مشترکی ثبت نشده است.
}}

{{#set:کد واحد={{{کد واحد|}}}|نام واحد={{{نام واحد|{{PAGENAME}}}}}|نوع واحد={{{نوع واحد|}}}|واحد بالادست={{{واحد بالادست|}}}|مدیر واحد={{{مدیر واحد|}}}|وضعیت واحد={{{وضعیت واحد|فعال}}}|شرح مأموریت={{{شرح مأموریت|}}}|کد استاندارد DED={{{کد استاندارد DED|}}}|حوزه واحد={{{حوزه واحد|داخلی}}}}}
</div>
[[رده:واحدهای سازمانی]]
</includeonly>"""

FORM_ORG = """<noinclude>
این فرم برای ایجاد یا ویرایش واحدهای سازمانی (داخلی و مرجع بیرونی) استفاده می‌شود.
{{#forminput:form=واحد سازمانی|button text=ایجاد/ویرایش واحد سازمانی|autocomplete on category=واحدهای سازمانی|placeholder=نام واحد را وارد کنید}}
</noinclude><includeonly>
{{{info|page name=<واحد سازمانی[نام واحد]>|create title=ایجاد واحد سازمانی|edit title=ویرایش واحد سازمانی}}}
{{{for template|واحد سازمانی}}}
{| class="formtable"
! کد واحد: || {{{field|کد واحد|size=25|placeholder=HAFA-EDU-ORG-001}}}
|-
! نام واحد: || {{{field|نام واحد|mandatory|size=60}}}
|-
! نوع واحد: || {{{field|نوع واحد|input type=dropdown|mandatory|values=دانشگاه,معاونت,مدیریت,مرکز,دفتر,اداره,گروه,کمیته,دانشکده,پردیس,مؤسسه آموزشی,مجتمع آموزش عالی,مؤسسه طرف تعهد,مرکز تحقیقاتی,مرکز رشد,پژوهشکده}}}
|-
! حوزه واحد: || {{{field|حوزه واحد|input type=dropdown|values=داخلی,مرجع بیرونی|default=داخلی}}}
|-
! کد استاندارد DED: || {{{field|کد استاندارد DED|size=25|placeholder=کد مرجع در کدینگ استاندارد آموزش}}}
|-
! واحد بالادست: || {{{field|واحد بالادست|input type=combobox|values from category=واحدهای سازمانی}}}
|-
! مدیر واحد: || {{{field|مدیر واحد|input type=combobox|values from category=افراد}}}
|-
! وضعیت واحد: || {{{field|وضعیت واحد|input type=dropdown|values=فعال,ادغام‌شده,منحل‌شده|default=فعال}}}
|-
! شرح مأموریت: || {{{field|شرح مأموریت|input type=textarea|rows=4|cols=70}}}
|}
{{{end template}}}
{{{standard input|save|label=ذخیره واحد}}} {{{standard input|cancel|label=انصراف}}}
</includeonly>"""

SUBCATS = [
    ("رده:دانشگاه‌ها", "دانشگاه‌ها و مؤسسات دانشگاهی ثبت‌شده به‌عنوان واحد مرجع بیرونی از کدینگ استاندارد آموزش."),
    ("رده:دانشکده‌ها و پردیس‌ها", "دانشکده‌ها و پردیس‌های خودگردان ثبت‌شده از کدینگ استاندارد آموزش (جدول ۶)."),
    ("رده:مؤسسات آموزش عالی", "مؤسسات آموزش عالی سلامت و مؤسسات مستقل ثبت‌شده از کدینگ استاندارد آموزش."),
    ("رده:مجتمع‌های آموزش عالی", "مجتمع‌های آموزش عالی علوم پزشکی ثبت‌شده از کدینگ استاندارد آموزش."),
    ("رده:مؤسسات طرف تعهد", "مؤسسات طرف تعهد ثبت‌شده از کدینگ استاندارد آموزش (جدول ۹)."),
    ("رده:مراکز تحقیقاتی", "مراکز تحقیقاتی ثبت‌شده از کدینگ استاندارد آموزش (جدول ۳۴)."),
    ("رده:مراکز رشد", "مراکز رشد ثبت‌شده از کدینگ استاندارد آموزش (جدول ۳۵)."),
    ("رده:پژوهشکده‌ها", "پژوهشکده‌های ثبت‌شده از کدینگ استاندارد آموزش (جدول ۳۶)."),
]


def build_patch_xml():
    pages = [
        xml_page("الگو:واحد سازمانی", 10, TPL_ORG),
        xml_page("فرم:واحد سازمانی", 106, FORM_ORG),
        xml_page("Property:کد استاندارد DED", 102, "[[Has type::Text]]\n[[Display title::کد استاندارد DED]]\n"),
        xml_page("Property:حوزه واحد", 102, "[[Has type::Text]]\n[[Display title::حوزه واحد]]\n"),
    ]
    for title, desc in SUBCATS:
        pages.append(xml_page(title, 14, f"{desc}\n\n[[رده:واحدهای سازمانی]]\n"))
    return xml_doc(pages)


# ---------------------------------------------------------------- اکسل
FONT_TITLE = Font(name="B Nazanin", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="B Nazanin", size=11, color="1F1F1F")
FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_ALT = PatternFill("solid", fgColor="DEEBF7")
THIN = Side(style="thin", color="9BC2E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def style_sheet(ws, title, headers, rows, widths):
    ws.sheet_view.rightToLeft = True
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = CENTER
    ws.row_dimensions[1].height = 26
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[2].height = 22
    for idx, row in enumerate(rows):
        r = 3 + idx
        for i, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=i, value=v)
            c.font = FONT_BODY; c.alignment = RIGHT; c.border = BORDER
            if idx % 2 == 1:
                c.fill = FILL_ALT
    ws.freeze_panes = "A3"


def build_terms_xlsx(terms):
    wb = Workbook(); ws = wb.active; ws.title = "واژگان حافا"
    headers = ["کد واژه", "نام واژه", "دسته", "نوع", "کد استاندارد DED", "تعریف",
               "منبع تعریف", "وضعیت واژه", "جدید (DED)", "قالب مقصد", "عنوان صفحه ویکی"]
    rows = [[t["کد واژه"], t["نام واژه"], t["دسته"], t["نوع"], t["کد DED"], t["تعریف"],
             t["منبع تعریف"], t["وضعیت"], t["جدید"], t["قالب"], t["عنوان صفحه"]] for t in terms]
    style_sheet(ws, "واژگان حافا — استخراج‌شده از کدینگ استاندارد آموزش (DED) — مدل تفکیک‌شده", headers, rows,
                [18, 32, 28, 16, 16, 70, 30, 14, 10, 12, 32])
    wb.save(XLSX_TERMS_FA)
    wb.save(XLSX_TERMS_EN)
    return len(terms)


def build_orgs_xlsx(orgs, ostan_rows, stats):
    wb = Workbook()
    ws = wb.active; ws.title = "واحدهای سازمانی"
    headers = ["حوزه واحد", "کد واحد (HAFA)", "نام واحد", "نوع واحد", "کد استاندارد DED",
               "دسته DED", "جدول DED", "وضعیت واحد", "جدید (DED)", "عنوان صفحه ویکی"]
    rows = [[o["حوزه"], o["کد واحد"], o["نام واحد"], o["نوع واحد"], o["کد DED"], o["دسته"],
             o["جدول"], "فعال", o["جدید"], o["عنوان صفحه"]] for o in orgs]
    style_sheet(ws, "واحدهای سازمانی حافا — داخلی (مبنای شروع) + مراجع بیرونی از کدینگ استاندارد آموزش (DED)",
                headers, rows, [12, 20, 45, 18, 16, 28, 12, 12, 10, 45])

    ws2 = wb.create_sheet("دامنه استان و شهر")
    style_sheet(ws2, "دامنه مقادیر عناصر داده «کد استان» و «کد شهرستان» — کدینگ استاندارد آموزش (DED)",
                ["کد استان", "استان", "کد شهر", "شهر"],
                [[r["کد استان"], r["استان"], r["کد شهر"], r["شهر"]] for r in ostan_rows],
                [12, 24, 12, 32])

    ws3 = wb.create_sheet("خلاصه تفکیک مدل")
    rows3 = [[cat, ("واحد سازمانی (مرجع بیرونی)" if cat in ORG_CATS else
                    ("عنصر داده (دامنه مقادیر)" if cat == OSTAN_CAT else "واژه")), n]
             for cat, n in stats.items()]
    style_sheet(ws3, "تفکیک دسته‌های DED در مدل داده حافا", ["دسته DED", "قالب مقصد", "تعداد رکورد خام"],
                rows3, [34, 30, 14])
    for p in (XLSX_ORGS_FA, XLSX_ORGS_EN):
        wb.save(p)
    return len(orgs)


# ---------------------------------------------------------------- main
def main():
    tables = extract_tables()
    terms, orgs_ext, ostan_rows, stats = build_all(tables)

    # یکتاسازی عنوان صفحات (سراسری بین واژه‌ها و واحدها به تفکیک فضای‌نام)
    used = {}
    for t in terms:
        title = t["نام واژه"]
        if title in used:
            used[title] += 1
            title = f"{t['نام واژه']} ({t['دسته']})"
            n = 1
            while title in used:
                n += 1
                title = f"{t['نام واژه']} ({t['دسته']} {n})"
        used[title] = 1
        t["عنوان صفحه"] = title
    for i, t in enumerate(terms, start=1):
        t["کد واژه"] = f"HAFA-EDU-TERM-{i:03d}"

    internal = build_internal(orgs_ext)
    orgs = internal + orgs_ext

    used = {}
    for o in orgs:
        title = o["نام واحد"]
        if title in used:
            used[title] += 1
            title = f"{o['نام واحد']} ({o['دسته']})"
            n = 1
            while title in used:
                n += 1
                title = f"{o['نام واحد']} ({o['دسته']} {n})"
        used[title] = 1
        o["عنوان صفحه"] = title

    # XML واژه‌ها
    with open(XML_TERMS, "w", encoding="utf-8") as f:
        f.write(xml_doc([xml_page(f"واژه:{t['عنوان صفحه']}", NS_TERM, term_wikitext(t)) for t in terms]))

    # XML واحدها + سامانه DED + عناصر داده + صفحه مرجع
    pages = [xml_page(o["عنوان صفحه"], 0, org_wikitext(o)) for o in orgs]
    pages.append(xml_page("سامانه:کدینگ استاندارد آموزش (DED)", NS_SYS, sys_wikitext()))
    el_ostan = data_element_wikitext(
        "HAFA-EDU-DATA-901", "کد استان",
        "شناسه عددی استاندارد استان مطابق جدول استان‌های کدینگ استاندارد آموزش (DED)؛ "
        "برای ثبت محل جغرافیایی در داده‌های آموزشی (مانند محل دانشگاه، محل تولد، محل خدمت).",
        "کد استاندارد (عدد صحیح)", "عدد صحیح مطابق کدینگ DED",
        f"۳۱ استان استاندارد DED — فهرست کامل: [[حافا:فهرست مرجع استان‌ها و شهرستان‌ها]]")
    el_shahr = data_element_wikitext(
        "HAFA-EDU-DATA-902", "کد شهرستان",
        "شناسه عددی استاندارد شهرستان مطابق جدول شهرهای کدینگ استاندارد آموزش (DED)؛ "
        "مقدار مجاز آن در ترکیب با کد استان معتبر است.",
        "کد استاندارد (عدد صحیح)", "عدد صحیح مطابق کدینگ DED",
        f"شهرستان‌های استاندارد DED به تفکیک استان — فهرست کامل: [[حافا:فهرست مرجع استان‌ها و شهرستان‌ها]]")
    pages.append(xml_page("داده:کد استان", NS_DATA, el_ostan))
    pages.append(xml_page("داده:کد شهرستان", NS_DATA, el_shahr))
    pages.append(xml_page("حافا:فهرست مرجع استان‌ها و شهرستان‌ها", NS_HAFA, ostan_ref_wikitext(ostan_rows)))
    with open(XML_ORGS, "w", encoding="utf-8") as f:
        f.write(xml_doc(pages))

    # پچ تعاریف
    with open(XML_PATCH, "w", encoding="utf-8") as f:
        f.write(build_patch_xml())

    # اکسل‌ها
    n_t = build_terms_xlsx(terms)
    n_o = build_orgs_xlsx(orgs, ostan_rows, stats)

    print(f"واژه‌ها: {n_t} صفحه")
    print(f"واحدهای سازمانی: {n_o} صفحه (داخلی: {len(internal)}، مرجع بیرونی: {len(orgs_ext)})")
    print(f"ردیف‌های استان/شهر (دامنه مقادیر): {len(ostan_rows)}")
    print("— تفکیک دسته‌ها:")
    for cat, n in stats.items():
        dest = "واحد سازمانی" if cat in ORG_CATS else ("عنصر داده" if cat == OSTAN_CAT else "واژه")
        print(f"   {cat}: {n} → {dest}")


if __name__ == "__main__":
    main()
