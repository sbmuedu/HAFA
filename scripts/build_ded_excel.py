# -*- coding: utf-8 -*-
"""
استخراج اطلاعات صفحات «کدینگ استاندارد آموزش» (DED) وزارت بهداشت،
دسته‌بندی و نگارش آن‌ها در یک فایل اکسل — برای استفاده در تهیه صفحات واژگان.
منبع: فایل‌های HTML استخراج‌شده از ded_data.zip (در پوشه ded_extract/)
"""
import os, re, glob
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = "/home/user/ded_extract"
OUT = "/home/user/HAFA/کدینگ_استاندارد_آموزش_دسته‌بندی‌شده.xlsx"

# ---------- استایل ----------
FONT_TITLE = Font(name="B Nazanin", size=14, bold=True, color="FFFFFF")
FONT_HEADER = Font(name="B Nazanin", size=11, bold=True, color="FFFFFF")
FONT_BODY = Font(name="B Nazanin", size=11, color="1F1F1F")
FONT_BOLD = Font(name="B Nazanin", size=11, bold=True, color="1F1F1F")
FILL_TITLE = PatternFill("solid", fgColor="1F4E79")
FILL_HEADER = PatternFill("solid", fgColor="2E75B6")
FILL_ALT = PatternFill("solid", fgColor="DEEBF7")
FILL_SECTION = PatternFill("solid", fgColor="FCE4D6")
THIN = Side(style="thin", color="9BC2E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
RIGHT_TOP = Alignment(horizontal="right", vertical="top", wrap_text=True)


def style_sheet(ws, widths):
    ws.sheet_view.rightToLeft = True
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, text, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font = FONT_TITLE; c.fill = FILL_TITLE; c.alignment = CENTER
    ws.row_dimensions[1].height = 26


def header_row(ws, row, headers):
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER; c.fill = FILL_HEADER; c.alignment = CENTER; c.border = BORDER
    ws.row_dimensions[row].height = 22


def data_row(ws, row, values, alt=False):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = FONT_BODY; c.alignment = RIGHT; c.border = BORDER
        if alt:
            c.fill = FILL_ALT


def note_row(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_BOLD; c.fill = FILL_SECTION; c.alignment = RIGHT_TOP
    for i in range(1, ncols + 1):
        ws.cell(row=row, column=i).border = BORDER


# ---------- ابزار استخراج ----------
def load_rows(fname):
    """بازگرداندن سطرهای همه جدول‌ها (با حفظ خانه‌های خالی)."""
    soup = BeautifulSoup(open(os.path.join(SRC, fname), encoding="utf-8", errors="replace").read(), "lxml")
    rows = []
    for t in soup.find_all("table"):
        for tr in t.find_all("tr"):
            rows.append([c.get_text(" ", strip=True) for c in tr.find_all(["td", "th"])])
    return rows


def clean(s):
    return re.sub(r"\s+", " ", s).strip()


def find_table_title(rows):
    """عنوان جدول (مثل «جدول (1): ...»)."""
    for r in rows:
        if len(r) == 1 and "جدول" in r[0]:
            return clean(r[0])
    return ""


def collect_notes(rows):
    """متن‌های توضیحی بلند قبل/بعد از جدول (یادداشت‌های حاکمیتی)."""
    notes = []
    for r in rows:
        if len(r) == 1:
            t = clean(r[0])
            if len(t) > 40 and "جدول" not in t:
                notes.append(t)
    return notes


def parse_simple(rows, title_idx, code_idx, desc_idx=None, row_idx=1):
    """جدول ساده: [.., ردیف, عنوان, کد, (توضیحات)]. خروجی: لیست دیکشنری."""
    recs = []
    for r in rows:
        if len(r) <= max(title_idx, code_idx):
            continue
        title = clean(r[title_idx]) if title_idx < len(r) else ""
        code = clean(r[code_idx]) if code_idx < len(r) else ""
        desc = clean(r[desc_idx]) if (desc_idx is not None and desc_idx < len(r)) else ""
        rid = clean(r[row_idx]) if row_idx < len(r) else ""
        if not title and not code:
            continue
        if "کد" in code or title == "ردیف" or "ردیف" in title:
            continue  # سطر سرستون
        # تشخیص پرچم «جدید»
        flag = ""
        cells = [c for c in r]
        if "جدید" in cells:
            flag = "جدید"
            cells = [c for c in cells if c != "جدید"]
        recs.append({"ردیف": rid, "کد": code, "عنوان": title, "توضیحات": desc, "جدید": flag})
    return recs


def parse_azad(rows):
    """دانشگاه آزاد: دو جفت (نام، کد) در یک سطر."""
    recs = []
    for r in rows:
        if len(r) < 6:
            continue
        pairs = [(1, 2), (4, 5)]
        for ni, ci in pairs:
            name = clean(r[ni]); code = clean(r[ci])
            if not name or not code or name in ("نام دانشگاه",):
                continue
            recs.append({"ردیف": "", "کد": code, "عنوان": name, "توضیحات": "", "جدید": ""})
    return recs


def parse_ostan(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        k_o = clean(r[1]); ostan = clean(r[2]); k_sh = clean(r[3]); shahr = clean(r[4])
        if not ostan and not shahr:
            continue
        if ostan == "استان":
            continue
        recs.append({"کد استان": k_o, "استان": ostan, "کد شهر": k_sh, "شهر": shahr})
    return recs


def parse_gerayesh(rows):
    recs = []
    for r in rows:
        if len(r) < 5:
            continue
        rid, reshte, maghta, gerayesh, code = (clean(x) for x in r[:5])
        if reshte == "عنوان رشته":
            continue
        recs.append({"ردیف": rid, "رشته": reshte, "مقطع": maghta, "گرایش": gerayesh, "کد": code})
    return recs


# ---------- مشخصات جداول کدینگ ----------
# (کلید، دسته، شماره جدول، فایل، نوع، ...)
CODE_TABLES = [
    # (دسته, شماره, فایل, parser_kind, title_idx, code_idx, desc_idx)
    ("مقاطع تحصیلی", "جدول 1", "Garde_ddf4520e04.html", "simple", 2, 3, 4),
    ("جنسیت", "جدول 13", "جنسیت_08631504e1.html", "simple", 2, 3, None),
    ("وضعیت تأهل", "جدول 11", "وضعیت-تاهل_875556740a.html", "simple", 2, 3, None),
    ("وضعیت دانشجو", "جدول 12", "وضعیت-دانشجو_eea442c453.html", "simple", 2, 3, None),
    ("وضعیت دوره", "جدول 14", "وضعیت-دوره_e46995bd92.html", "simple", 2, 3, 4),
    ("وضعیت نظام وظیفه", "جدول 4", "وضعیت-نظام-وظیفه_03a48d9e8e.html", "simple", 2, 3, 4),
    ("نوع مقاطع", "جدول 15", "نوع-مقاطع_3301a51615.html", "simple", 2, 3, None),
    ("نوع تعهد", "جدول 10", "نوع-تعهد_f954c989c5.html", "simple", 2, 3, 4),
    ("دانشکده‌ها و پردیس‌ها", "جدول 6", "دانشکده‌-ها-و-پردیس-های-خودگردان_7ddaee6716.html", "simple", 2, 3, None),
    ("دانشگاه و موسسات مستقل و وابسته", "جدول 1-5", "دانشگاه-و-موسسات-پزشکی-مستقل-و-وابسته_98fb4ec5e3.html", "simple", 2, 3, None),
    ("دانشگاه‌های آزاد و غیرانتفاعی", "جدول 2-5", "دانشگاه‌های-آزاد-و-غیردولتی–غیرانتفاعی_d45da6185c.html", "azad", None, None, None),
    ("دانشگاه‌ها و مراکز غیروابسته", "جدول 3-5", "دانشگاه‌-ها-و-مراکز-آموزشی-غیروابسته_fa1dbc727c.html", "simple", 2, 3, None),
    ("موسسات/مجتمع آموزش عالی سلامت", "جدول 7", "موسسه-و-مجتمع-آموزش-عالی-سلامت_aebccdc3de.html", "simple", 2, 3, None),
    ("مجتمع آموزش عالی علوم پزشکی", "جدول 8", "مجتمع-آموزش-عالی-علوم-پزشکی_7135f96de0.html", "simple", 2, 3, None),
    ("موسسات طرف تعهد", "جدول 9", "موسسات-طرف-تعهد_54b201212f.html", "simple", 2, 3, None),
    ("مراکز تحقیقاتی", "جدول 34", "مراکز-تحقیقاتی_ff2cf137f9.html", "simple", 2, 3, None),
    ("مراکز رشد", "جدول 35", "مراکز-رشد_189461942c.html", "simple", 2, 3, None),
    ("پژوهشکده‌ها", "جدول 36", "پژوهشکده-ها_f3ed20cb79.html", "simple", 2, 3, None),
    ("گرایش‌های رشته‌ها", "جدول 3", "گرایشهای-استاندارد-رشته-ها-در-مقاطع-تحصیلی_a351189a42.html", "gerayesh", None, None, None),
    ("استان‌ها و شهرها", "جدول (استان/شهر)", "Ostan_a1d830cb84.html", "ostan", None, None, None),
]


def extract_code_table(spec):
    cat, tno, fname, kind, ti, ci, di = spec
    rows = load_rows(fname)
    title = find_table_title(rows)
    notes = collect_notes(rows)
    if kind == "simple":
        recs = parse_simple(rows, ti, ci, di)
    elif kind == "azad":
        recs = parse_azad(rows)
    elif kind == "ostan":
        recs = parse_ostan(rows)
    elif kind == "gerayesh":
        recs = parse_gerayesh(rows)
    else:
        recs = []
    return {"category": cat, "table_no": tno, "file": fname,
            "title": title, "notes": notes, "records": recs}


# ============================================================
# اجرای استخراج
# ============================================================
data = {}
for spec in CODE_TABLES:
    d = extract_code_table(spec)
    data[spec[0]] = d

# --- رشته‌ها: فهرست مقاطع ---
reshteha_rows = load_rows("Reshteha_535de1b616.html")
reshteha_note = collect_notes(reshteha_rows)
reshteha_list = []
for r in reshteha_rows:
    if len(r) >= 1:
        t = clean(r[0])
        if t.startswith("-"):
            reshteha_list.append(t.lstrip("- ").strip())

# --- موارد جدید کدینگ (changelog) ---
coding_rows = load_rows("Coding_0c52a7c893.html")
coding_items = []
cur_section = ""
for r in coding_rows:
    if len(r) < 2:
        continue
    c0 = clean(r[0]); c1 = clean(r[1])
    if c0.startswith("-"):
        cur_section = c0.lstrip("- ").strip()
        if c1:
            coding_items.append({"بخش": cur_section, "مورد": c1, "کد": ""})
    else:
        if c1:
            m = re.search(r"کد[^0-9]*?(\d[\d\s]*)\s*$", c1)
            code = m.group(1).strip() if m else ""
            coding_items.append({"بخش": cur_section, "مورد": c1, "کد": code})

# ============================================================
# ساخت اکسل
# ============================================================
wb = Workbook()

# ---------- 1) فهرست صفحات ----------
ws = wb.active
ws.title = "فهرست صفحات"
style_sheet(ws, [5, 34, 26, 22, 18, 60])
title_row(ws, "فهرست و دسته‌بندی صفحات کدینگ استاندارد آموزش (DED)", 6)
header_row(ws, 2, ["ردیف", "عنوان/دسته", "فایل منبع", "نوع محتوا", "شماره جدول", "یادداشت/توضیحات"])

index_rows = [
    ("صفحه اصلی معاونت", "صفحه-اصلی-معاونت_ee2154a111.html", "متن توصیفی", "—", "صفحه اصلی سایت معاونت آموزشی (اخبار و اطلاعیه‌ها)."),
    ("index (آرشیو)", "index_22118a3ae2.html", "متن توصیفی", "—", "آرشیو اخبار و رویدادهای معاونت اجرایی."),
    ("درباره معاونت", "درباره-معاونت_cdd266941d.html", "متن توصیفی", "—", "صفحه «درباره معاونت» (بدون محتوای متنی اصلی)."),
    ("معرفی", "معرفی_8bff32b4e9.html", "متن توصیفی", "—", "معرفی معاونت اجرایی، واحدهای تحت پوشش، مدیر و مسئول دفتر."),
    ("معاونین گذشته", "معاونین-گذشته_5af00e8168.html", "متن توصیفی", "—", "فهرست معاونین اجرایی گذشته با مدرک و دوره تصدی."),
    ("تماس با ما", "تماس-با-ما_8341d8f57a.html", "متن توصیفی", "—", "آدرس، تلفن و نمابر معاونت."),
    ("مقدمه", "مقدمه_de94b6b3c4.html", "متن توصیفی", "—", "مقدمه مجموعه کدینگ آموزش + فهرست همکاران تدوین‌کننده."),
    ("واحد آمار و فناوری اطلاعات (IT)", "IT_d83b52f769.html", "متن توصیفی", "—", "صفحه واحد آمار و فناوری اطلاعات (بدون جدول داده)."),
    ("موارد جدید کدینگ", "Coding_0c52a7c893.html", "جدول کدینگ (تغییرات)", "—", "فهرست موارد جدید/اصلاح‌شده در همه جداول کدینگ."),
    ("رشته‌ها (فهرست مقاطع)", "Reshteha_535de1b616.html", "فهرست", "—", "فهرست مقاطع دارای جدول رشته‌ها (هر مقطع صفحه جداگانه دارد)."),
]
r = 3
for i, (a, b, c, d, e) in enumerate(index_rows, start=1):
    data_row(ws, r, [i, a, b, c, d, e], alt=(i % 2 == 0))
    r += 1

# جداول کدینگ در فهرست
for spec in CODE_TABLES:
    cat, tno, fname, *_ = spec
    d = data[cat]
    nrec = len(d["records"])
    note = (d["notes"][0][:70] + "…") if d["notes"] else ""
    data_row(ws, r, [len(index_rows) + 1, cat, fname, "جدول کد استاندارد", tno, f"{nrec} قلم | {note}"], alt=(r % 2 == 1))
    r += 1

# ---------- 2) صفحات توصیفی ----------
ws = wb.create_sheet("صفحات توصیفی")
style_sheet(ws, [24, 46, 60])
title_row(ws, "صفحات توصیفی سایت معاونت (برای واژگان معرفی و پیشینه)", 3)
header_row(ws, 2, ["صفحه", "قلم/فرد", "توضیحات"])

r = 3
note_row(ws, r, "مقدمه", 3); r += 1
intro_rows = load_rows("مقدمه_de94b6b3c4.html")
for rr in intro_rows:
    if len(rr) >= 1 and clean(rr[0]):
        txt = clean(rr[0])
        if len(txt) > 60:
            data_row(ws, r, ["مقدمه", "متن مقدمه", txt], alt=False)
            r += 1
        elif ":" in txt or "دكتر" in txt or "مهندس" in txt or "سركار" in txt or "جناب" in txt:
            data_row(ws, r, ["مقدمه", "همکار", txt], alt=False)
            r += 1

note_row(ws, r, "معرفی معاونت اجرایی", 3); r += 1
moarefi = load_rows("معرفی_8bff32b4e9.html")
for rr in moarefi:
    if len(rr) >= 1 and clean(rr[0]):
        txt = clean(rr[0])
        if txt.startswith("معاونت اجرایی"):
            data_row(ws, r, ["معرفی", "رسالت", txt], alt=False)
        elif txt.startswith("واحدهای"):
            data_row(ws, r, ["معرفی", "واحدهای تحت پوشش", txt.replace("واحدهای تحت پوشش این حوزه شامل موارد ذیل است:", "")], alt=False)
        elif txt.startswith("مدیر"):
            data_row(ws, r, ["معرفی", "مدیر", txt], alt=False)
        elif txt.startswith("مسئول دفتر"):
            data_row(ws, r, ["معرفی", "مسئول دفتر", txt], alt=False)
        r += 1

note_row(ws, r, "معاونین گذشته", 3); r += 1
vp = [
    ("پرویز اسفندیاری", "کارشناسی ارشد مدیریت دولتی", ""),
    ("دکتر سید محمد پورحسینی", "Ph.D ژنتیک پزشکی", "91/12/7 - 92/9/25"),
    ("دکتر محمدحسین افجه‌ای", "Ph.D فیزیوتراپی", ""),
    ("دکتر حسن ابوالقاسم گرجی", "Ph.D مدیریت خدمات بهداشتی درمانی", "92/2/25 - 94/3/20"),
    ("دکتر امیراحمد اخوان", "Ph.D حشره‌شناسی پزشکی و مبارزه با ناقلین", "94/02/03 - 1400/07/24"),
]
for name, degree, period in vp:
    data_row(ws, r, ["معاونین گذشته", name, f"{degree}" + (f" | دوره تصدی: {period}" if period else "")], alt=(r % 2 == 1))
    r += 1

note_row(ws, r, "تماس با ما", 3); r += 1
contact = [
    ("آدرس", "ایران، تهران، شهرک غرب، ایوانک شرق، بین فلامک و زرافشان، ساختمان وزارت بهداشت، بلوک C، طبقه سیزدهم"),
    ("تلفن", "88364235 - 81452923"),
    ("نمابر", "88363860"),
]
for k, v in contact:
    data_row(ws, r, ["تماس با ما", k, v], alt=(r % 2 == 1))
    r += 1

note_row(ws, r, "درباره معاونت / صفحه اصلی / آرشیو", 3); r += 1
data_row(ws, r, ["درباره معاونت", "محتوا", "صفحه فاقد متن اصلی است (فقط ناوبری و پاورقی)."], alt=False)
r += 1
data_row(ws, r, ["صفحه اصلی", "محتوا", "اخبار و اطلاعیه‌های معاونت آموزشی (آزمون‌ها، بازدیدها، اطلاعیه‌ها)."], alt=False)
r += 1
data_row(ws, r, ["آرشیو (index)", "محتوا", "آرشیو رویدادها و گزارش‌های معاونت اجرایی."], alt=False)

# ---------- جداول کدینگ ----------
def write_simple_sheet(cat, headers, rec_map):
    d = data[cat]
    ws = wb.create_sheet(cat)
    ncols = len(headers)
    style_sheet(ws, [14, 16, 46, 60, 14])
    title_row(ws, f"{d['title'] or cat}", ncols)
    header_row(ws, 2, headers)
    rr = 3
    for i, rec in enumerate(d["records"]):
        vals = [rec_map(rec)[h] for h in headers]
        data_row(ws, rr, vals, alt=(i % 2 == 1))
        rr += 1
    # یادداشت‌ها
    for n in d["notes"]:
        rr += 1
        note_row(ws, rr, n, ncols)
    return ws


def sheet_name(s):
    return re.sub(r"[/\\*?\[\]:]", "-", s)[:31]


def write_code_sheet(cat):
    d = data[cat]
    headers = ["ردیف", "کد استاندارد", "عنوان/نام", "توضیحات", "وضعیت"]
    ws = wb.create_sheet(sheet_name(cat))
    style_sheet(ws, [10, 16, 44, 60, 12])
    title_row(ws, d["title"] or cat, len(headers))
    header_row(ws, 2, headers)
    rr = 3
    for i, rec in enumerate(d["records"]):
        vals = [rec.get("ردیف", ""), rec.get("کد", ""), rec.get("عنوان", ""),
                rec.get("توضیحات", ""), rec.get("جدید", "")]
        data_row(ws, rr, vals, alt=(i % 2 == 1))
        rr += 1
    for n in d["notes"]:
        rr += 1
        note_row(ws, rr, n, len(headers))
    return ws


# شیت‌های جدول ساده
simple_cats = [
    "مقاطع تحصیلی", "جنسیت", "وضعیت تأهل", "وضعیت دانشجو", "وضعیت دوره",
    "وضعیت نظام وظیفه", "نوع مقاطع", "نوع تعهد", "دانشکده‌ها و پردیس‌ها",
    "دانشگاه و موسسات مستقل و وابسته", "دانشگاه‌ها و مراکز غیروابسته",
    "موسسات/مجتمع آموزش عالی سلامت", "مجتمع آموزش عالی علوم پزشکی",
    "موسسات طرف تعهد", "مراکز تحقیقاتی", "مراکز رشد", "پژوهشکده‌ها",
]
for cat in simple_cats:
    write_code_sheet(cat)

# دانشگاه‌های آزاد (ساختار ساده)
write_code_sheet("دانشگاه‌های آزاد و غیرانتفاعی")

# گرایش‌ها
d = data["گرایش‌های رشته‌ها"]
ws = wb.create_sheet(sheet_name("گرایش‌های رشته‌ها"))
style_sheet(ws, [10, 34, 22, 34, 16])
title_row(ws, d["title"], 5)
header_row(ws, 2, ["ردیف", "عنوان رشته", "مقطع", "گرایش", "کد استاندارد گرایش"])
rr = 3
for i, rec in enumerate(d["records"]):
    data_row(ws, rr, [rec["ردیف"], rec["رشته"], rec["مقطع"], rec["گرایش"], rec["کد"]], alt=(i % 2 == 1))
    rr += 1

# استان‌ها و شهرها
d = data["استان‌ها و شهرها"]
ws = wb.create_sheet(sheet_name("استان‌ها و شهرها"))
style_sheet(ws, [10, 16, 30, 16, 30])
title_row(ws, d["title"], 5)
header_row(ws, 2, ["ردیف", "کد استان", "استان", "کد شهر", "شهر"])
rr = 3
for i, rec in enumerate(d["records"]):
    data_row(ws, rr, [i, rec["کد استان"], rec["استان"], rec["کد شهر"], rec["شهر"]], alt=(i % 2 == 1))
    rr += 1

# رشته‌ها (فهرست مقاطع)
ws = wb.create_sheet(sheet_name("رشته‌ها - فهرست مقاطع"))
style_sheet(ws, [6, 60])
title_row(ws, "رشته‌ها و کدهای استاندارد — فهرست مقاطع", 2)
header_row(ws, 2, ["ردیف", "مقطع"])
rr = 3
for i, m in enumerate(reshteha_list, start=1):
    data_row(ws, rr, [i, m], alt=(i % 2 == 1))
    rr += 1
for n in reshteha_note:
    rr += 1
    note_row(ws, rr, n, 2)

# موارد جدید کدینگ
ws = wb.create_sheet(sheet_name("موارد جدید کدینگ"))
style_sheet(ws, [6, 36, 60, 16])
title_row(ws, "موارد جدید و اصلاح‌شده کدینگ (Changelog)", 4)
header_row(ws, 2, ["ردیف", "بخش/جدول", "مورد جدید", "کد استاندارد"])
rr = 3
for i, it in enumerate(coding_items, start=1):
    data_row(ws, rr, [i, it["بخش"], it["مورد"], it["کد"]], alt=(i % 2 == 1))
    rr += 1

# ---------- شیت تجمیعی «همه اقلام» ----------
ws = wb.create_sheet("همه اقلام (تجمیع)")
style_sheet(ws, [6, 30, 16, 16, 44, 60, 12])
title_row(ws, "همه اقلام کدینگ استاندارد — تجمیع (برای تهیه صفحات واژگان)", 7)
header_row(ws, 2, ["ردیف", "دسته", "شماره جدول", "کد استاندارد", "عنوان/نام", "توضیحات", "وضعیت"])
rr = 3
idx = 0
for cat in CODE_TABLES:
    cat_name = cat[0]
    d = data[cat_name]
    if cat[3] == "ostan":
        for rec in d["records"]:
            idx += 1
            data_row(ws, rr, [idx, cat_name, d["table_no"], rec["کد استان"] + "/" + rec["کد شهر"],
                              rec["استان"] + " — " + rec["شهر"], "", ""], alt=(idx % 2 == 0))
            rr += 1
    elif cat[3] == "gerayesh":
        for rec in d["records"]:
            idx += 1
            data_row(ws, rr, [idx, cat_name, d["table_no"], rec["کد"],
                              f"{rec['رشته']} / {rec['مقطع']} / {rec['گرایش']}", "", ""], alt=(idx % 2 == 0))
            rr += 1
    else:
        for rec in d["records"]:
            idx += 1
            data_row(ws, rr, [idx, cat_name, d["table_no"], rec.get("کد", ""),
                              rec.get("عنوان", ""), rec.get("توضیحات", ""), rec.get("جدید", "")], alt=(idx % 2 == 0))
            rr += 1

wb.save(OUT)
print("Saved:", OUT)
print("Sheets:", wb.sheetnames)
print("Total records in master:", idx)
